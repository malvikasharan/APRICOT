#! /usr/bin/env python

import os
import sys
import csv
import shutil
csv.field_size_limit(sys.maxsize)
try:
    from openpyxl.workbook import Workbook
except ImportError:
    print('Python package openpyxl is missing. Please install/update.\n'
    'Please ignore if you chose the output format as HTML')

__description__ = '''This script is designed to convert the file types.
tab separated files to xlsx files'''
__author__ = "Malvika Sharan <malvika.sharan@uni-wuerzburg.de>"
__email__ = "malvika.sharan@uni-wuerzburg.de"


def csv_to_xlsx(inpath, outpath):
    '''Converts table to excel'''
    allowed_format_list = ['csv', 'tab', 'tsv']
    for csv_file in os.listdir(inpath):
        if os.path.isfile(inpath+'/'+csv_file):
            if csv_file.split('.')[1] in allowed_format_list:
                sheet_name = csv_file.split('.')[0]
                workbook = Workbook()
                xlsxsheet = workbook.create_sheet(0)
                for i, row in enumerate(
                        csv.reader(open(inpath+'/'+csv_file), delimiter="\t")):
                    for j, col in enumerate(row):
                        xlsxsheet.cell(row=i, column=j).value = col
                workbook.save(open(outpath+'/'+sheet_name+'.xlsx', 'wb'))
            else:
                shutil.copyfile(inpath+'/'+csv_file, outpath+'/'+csv_file)
        elif os.path.isdir(inpath+'/'+csv_file):
            if not os.path.exists(outpath+'/'+csv_file):
                os.mkdir(outpath+'/'+csv_file)
            csv_to_xlsx(inpath+'/'+csv_file, outpath+'/'+csv_file)

            
def csv_to_html(inpath, outpath):
    '''Convert table to HTML, line break 13-16'''
    allowed_format_list = ['csv', 'tab', 'tsv']
    for csv_file in os.listdir(inpath):
        if os.path.isfile(inpath+'/'+csv_file):
            if csv_file.split('.')[1] in allowed_format_list:
                with open(outpath+'/'+csv_file.split('.')[0]+'.html',"w") as outfile:
                    outfile.write('\n'.join(['<!DOCTYPE html>', '<html lang="en">', '<head>',
                    '<meta charset="utf-8">',
                    '<link rel="stylesheet" type="text/css" href="http://ajax.aspnetcdn.com/ajax/jquery.dataTables/1.9.4/css/jquery.dataTables.css">',
                    '<script type="text/javascript" charset="utf8" src="http://ajax.aspnetcdn.com/ajax/jQuery/jquery-1.8.2.min.js"></script>',
                    '<script type="text/javascript" charset="utf8" src="http://ajax.aspnetcdn.com/ajax/jquery.dataTables/1.9.4/jquery.dataTables.min.js"></script>',
                    '<script type="text/javascript" language="javascript" class="init">',
                    '$(document).ready(function () {',
                    '\t$("#subscriptionlist").dataTable();',
                    '});', '</script>', '<style type="text/CSS">',
                    'table, th, td {', '\tborder: 1px solid #E8E8E8;',
                    '\tborder-collapse: collapse;', '\tmin-width:150px;',
                    '\ttext-align:center', '}', '</style>',
                    '</head>', '<div id="subscriptionsList">',
                    '<h2>Source file: %s</h2>' % csv_file,
                    '<h3>Source path: %s</h3>' % inpath, '<table id="subscriptionlist">\n']))
                    row_num = 0
                    table_string = ""
                    csv_fh = csv.reader(open(inpath+'/'+csv_file), delimiter="\t")
                    for i, row in enumerate(csv_fh):
                        if 'annotation_scoring' in csv_file:
                            if 'annotation_scoring_of_selected_data_filter' in csv_file:
                                for idx in range(1, 3):
                                    if len(list(row[idx])) > 50:
                                        row[idx] = '<br>'.join(
                                            split_str(row[idx], 50))
                            else:
                                for idx in range(13, 17):
                                    if len(list(row[idx])) > 50:
                                        row[idx] = '<br>'.join(
                                            split_str(row[idx], 50))
                        if i == 0:
                            outfile.write("<thead>\n<tr>\n<th>")
                            outfile.write('</th><th>'.join(row)+'\n')
                            outfile.write("</th>\n</tr>\n</thead>\n<tbody>\n")
                        else:
                            outfile.write("<tr>\n<td>")
                            outfile.write('</td><td>'.join(row)+'\n')
                            outfile.write("</td>\n</tr>\n")
                    outfile.write(
                        '\n'.join(['</tbody>', '</table>', '</div>',
                                   '</html>']))
            else:
                shutil.copyfile(inpath+'/'+csv_file, outpath+'/'+csv_file)
        elif os.path.isdir(inpath+'/'+csv_file):
            if not os.path.exists(outpath+'/'+csv_file):
                os.mkdir(outpath+'/'+csv_file)
            csv_to_html(inpath+'/'+csv_file, outpath+'/'+csv_file)


def split_str(seq, chunk, skip_tail=False):
    lst = []
    if chunk <= len(seq):
        lst.extend([seq[:chunk]])
        lst.extend(split_str(seq[chunk:], chunk, skip_tail))
    elif not skip_tail and seq:
        lst.extend([seq])
    return lst
    
if __name__ == '__main__':
    inpath = sys.argv[1]
    outpath = sys.argv[2]
    csv_to_xlsx(inpath, outpath)
    csv_to_html(inpath, outpath)
